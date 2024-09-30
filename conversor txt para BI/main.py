import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserIconView
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import glob
import os

kivy.require('1.11.1')

class MyGrid(BoxLayout):

    def __init__(self, **kwargs):
        super(MyGrid, self).__init__(**kwargs)
        self.orientation = 'vertical'
        
        self.filechooser = FileChooserIconView()
        self.add_widget(self.filechooser)

        self.process_button = Button(text='Processar Arquivos')
        self.process_button.bind(on_press=self.process_files)
        self.add_widget(self.process_button)

        self.result_label = Label(text='')
        self.add_widget(self.result_label)

    def process_files(self, instance):
        selected_files = self.filechooser.selection
        if not selected_files:
            self.result_label.text = "Por favor, selecione os arquivos .txt."
            return

        output_folder = "output-convert"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        
        excel_file = os.path.join(output_folder, "saida.xlsx")

        txt_to_excel(selected_files, excel_file)

        self.result_label.text = f'Arquivos processados com sucesso. Salvo em: {excel_file}'

def txt_to_excel(txt_files, excel_file):
    data_frames = []

    for txt_file in txt_files:
        with open(txt_file, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        data = [line.strip().split() for line in lines]
        df = pd.DataFrame(data)
        data_frames.append(df)

    all_data = pd.concat(data_frames, ignore_index=True)
    all_data.to_excel(excel_file, index=False, header=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    font = Font(name='Calibri', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    wb.save(excel_file)

class MyApp(App):
    def build(self):
        return MyGrid()

if __name__ == "__main__":
    MyApp().run()
