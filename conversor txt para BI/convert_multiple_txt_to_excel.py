import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import glob
import os

def txt_to_excel(txt_files, excel_file):
    data_frames = []

    # Ler todos os arquivos .txt e concatená-los em um DataFrame
    for txt_file in txt_files:
        with open(txt_file, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        # Processar linhas do .txt (supondo que seja delimitado por espaços ou tabulações)
        data = [line.strip().split() for line in lines]

        # Criar um DataFrame do pandas
        df = pd.DataFrame(data)
        data_frames.append(df)

    # Concatenar todos os DataFrames
    all_data = pd.concat(data_frames, ignore_index=True)

    # Salvar o DataFrame concatenado em um arquivo Excel
    all_data.to_excel(excel_file, index=False, header=False)

    # Carregar o arquivo Excel para edição com openpyxl
    wb = load_workbook(excel_file)
    ws = wb.active

    # Definir a fonte personalizada
    font = Font(name='Calibri', size=11)  # Calibri é compatível com Power BI

    # Aplicar a fonte personalizada a todas as células
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    # Salvar o arquivo Excel com a formatação aplicada
    wb.save(excel_file)

    print(f'Arquivo Excel {excel_file} criado e formatado com sucesso!')

# Exemplo de uso
input_folder = 'input-convert'  # Diretório onde os arquivos .txt estão armazenados
output_folder = 'output-convert'  # Diretório onde o arquivo Excel será salvo
txt_files = glob.glob(os.path.join(input_folder, '*.txt'))  # Obter todos os arquivos .txt na pasta 'input-convert'

excel_file = input("Digite o nome do arquivo Excel de saída (inclua .xlsx no final): ")

# Verifica se o nome do arquivo inclui a extensão .xlsx
if not excel_file.endswith('.xlsx'):
    excel_file += '.xlsx'

# Adicionar o caminho da pasta 'output-convert' ao nome do arquivo de saída
excel_file = os.path.join(output_folder, excel_file)

txt_to_excel(txt_files, excel_file)
