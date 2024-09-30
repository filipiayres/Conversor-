import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def txt_to_excel(txt_file, excel_file):
    # Ler o arquivo .txt
    with open(txt_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # Processar linhas do .txt (supondo que seja delimitado por espaços ou tabulações)
    data = [line.strip().split() for line in lines]

    # Criar um DataFrame do pandas
    df = pd.DataFrame(data)

    # Salvar o DataFrame em um arquivo Excel
    df.to_excel(excel_file, index=False, header=False)

    # Carregar o arquivo Excel para edição com openpyxl
    wb = load_workbook(excel_file)
    ws = wb.active

    # Definir a fonte personalizada
    font = Font(name='Calibri', size=11)  # Calibri é compatível com Power BI

    # Aplicar a fonte personalizada a todas as células
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font

    # Salvar o arquivo Excel com a formatação aplicada
    wb.save(excel_file)

    print(f'Arquivo Excel {excel_file} criado e formatado com sucesso!')

# Exemplo de uso
txt_file = 'seu_arquivo.txt'  # Substitua pelo nome do seu arquivo .txt
excel_file = 'output.xlsx'    # O nome do arquivo Excel que será criado
txt_to_excel(txt_file, excel_file)
